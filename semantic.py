from openpyxl import load_workbook
import time

start = time.time() ## точка отсчета времени

# wb1 = load_workbook('D:/AdressClear/Src/rosreestr.xlsx')
# table_input = wb1.active

def get_kooperativ(input_adress):
    wb4 = load_workbook("./Src/kooperativ_list.xlsx")
    kooperativ_list = wb4.active
    kooperativ = "None"
    for j in range(1, kooperativ_list.max_row + 1):
        kooperativ = kooperativ_list[j][0].value
        index = input_adress.find(kooperativ)
        if index != -1:
            break
        else:
            kooperativ = "None"
    return kooperativ

def get_kvartal(input_adress):
    wb5 = load_workbook("./Src/kvartal.xlsx")
    kvartal_list = wb5.active
    kvartal_clear = "None"
    for j in range(1, kvartal_list.max_row + 1):
        kvartal = kvartal_list[j][0].value
        kvartal_clear = kvartal_list[j][1].value
        index = input_adress.find(kvartal)
        if index != -1:
            break
        else:
            kvartal_clear = "None"
    return kvartal_clear

def get_snt(input_adress):
    wb3 = load_workbook("./Src/snt_list.xlsx")
    snt_list = wb3.active
    snt = "None"
    for j in range(1, snt_list.max_row + 1):
        snt = snt_list[j][0].value
        index = input_adress.find(snt)
        if index != -1:
            break
        else:
            snt = "None"
    return snt

def get_street(input_adress):
    wb2 = load_workbook("./Src/street_list.xlsx")
    street_list = wb2.active
    name = ["40 лет Октября", "Бабушкина", "Володарского", "Гончарова", "Дарвина", "Дружинина",
            "Кольцова", "Короленко", "Котовского", "Левашова", "Ломоносова", "Менделеева", "Менжинского",
            "Пархоменко", "Редаковский", "Серова", "Сусанина", "Чернышевского"]
    for i in range(1, street_list.max_row + 1):
        ul = str(street_list[i][0].value)
        ul = ul.replace("-", '')
        ul = ul.replace(" ", '')
        ul_clear = str(street_list[i][1].value)
        # print("ul", ul)
        # print("input_adress", input_adress)
        index1 = input_adress.find(ul)
        # print("index1", index1)
        if index1 != -1:
            if ul_clear in name:
                input_adress = input_adress.replace(ul, '')
                input_adress = input_adress.lower()
                if (input_adress.find("пер") > 0) or\
                   (input_adress.find("пер.") > 0) or\
                   (input_adress.find("переулок") > 0):
                    type_street = str("переулок")
                elif (input_adress.find("проезд") > 0) or\
                     (input_adress.find("пр.") > 0) or\
                     (input_adress.find("пр-д") > 0) or \
                     (input_adress.find("пр") > 0):
                    type_street = str("проезд")
                elif (input_adress.find("подъем") > 0):
                    type_street = str("подъем")
                else:
                    type_street = str("улица")
            else:
                type_street = str(street_list[i][2].value)
            break
        else:
            ul_clear = "None"
            type_street = "None"
    return type_street, ul_clear
